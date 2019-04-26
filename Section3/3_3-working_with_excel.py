# -*- coding: utf-8 -*-
"""
Created on Thu Apr  4 13:04:22 2019

@author: Matt Macarty
"""
# module for processing Excel format files
# note you may need to install:
# from command line python -m pip install openpyxl
import openpyxl as oxl

# there are other Excel to Python libraries out there
# but they tend to work about the same way

excel = oxl.load_workbook('emails.xlsx')
data = excel['Sheet1']

# get last row
last_row = data.max_row

# create list of Names
emails =[]
names = []
for i in range(2, last_row +1):
    flag = data.cell(row = i, column = 3 ).value
    if flag != "N":
        names.append(data.cell(row = i, column = 1).value)
        

# create list of emails
for i in range(2, last_row +1):
    flag = data.cell(row = i, column = 3 ).value
    if flag != "N":
        emails.append(data.cell(row = i, column = 2).value)
        
        


